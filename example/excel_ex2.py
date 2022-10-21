# 2022.10 HDY
# 파워유저교육(파이썬) - 엑셀 다루기 예제

# pip install openpyxl

# Excel에 데이터 쓰기

from openpyxl import load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook(r'excel_ex1.xlsx', data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['Sheet']

# 셀 주소로 값 읽어오기
print(load_ws['A1'].value)          # 숫자

# 셀 좌표로 값 읽어오기
print(load_ws.cell(2, 2).value)     # 1

# 범위로 값 읽어오기
get_cells = load_ws['A2' : 'C2']    # 1, 2, 3
for row in get_cells:
    for cell in row:
        print(cell.value)

# 모든 행 단위로 읽기
for row in load_ws.rows:
    print(row)

# 모든 열 단위로 읽기
for col in load_ws.columns:
    print(col)

# 모든 행과 열 읽기
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)