# 2022.10 HDY
# 파워유저교육(파이썬) - 엑셀 다루기 예제

# pip install openpyxl

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 워크북 생성
wb = openpyxl.Workbook()

# 시트 활성화
ws = wb.worksheets[0] # wb.active 로 써도 됨

# 시트 이름 바꾸기
ws.title = 'worksheet'

# 새로운 시트 만들기
ws2 = wb.create_sheet('worksheet2')

# 두번째 worksheet2의 'B2'에 'cell' 쓰기
ws2['B2'] = 'cell'

# 워크북 filename 으로 저장(없으면 새로 생성)
filename = 'test.xlsx'
wb.save(filename)

# Cell 꾸미기

# 가운데 정렬
align_center = Alignment(horizontal='center', vertical='center')

# 글씨체 굵게
font_bold = Font(size=12, bold=True, color='0000FF') # 0000FF Blue

# Cell 색깔 채우기
fill_yellow = PatternFill('solid', fgColor='FFFF00')

# 테두리 선넣기
thin_border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

# 위의 4가지 서식 B2 Cell에 적용ㅇ시키기 + 값 'cell'에서 'text'로 바꾸기
ws['B2'].alignment = align_center
ws['B2'].font = font_bold
ws['B2'].fill = fill_yellow
ws['B2'].border = thin_border
ws['B2'].value = 'text'

# 워크북 저장
wb.save(filename)
